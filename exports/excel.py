import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EMPRESA_NOMBRE

AZUL      = "1E4E8C"
AZUL_CLAR = "DDEEFF"
GRIS      = "F2F2F2"
BLANCO    = "FFFFFF"

def _estilo_header(ws, fila, columnas):
    for col, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col, value=titulo)
        cell.font      = Font(bold=True, color=BLANCO, size=10)
        cell.fill      = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borde()

def _borde():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _titulo_reporte(ws, titulo):
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value     = EMPRESA_NOMBRE
    c.font      = Font(bold=True, size=13, color=AZUL)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    c2 = ws["A2"]
    c2.value     = titulo
    c2.font      = Font(bold=True, size=11)
    c2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

def exportar_ventas(datos: list, titulo: str = "Reporte de Ventas") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    _titulo_reporte(ws, titulo)

    columnas = [
        "Nº Doc", "Fecha", "Tipo Doc", "Tipo Venta",
        "Cliente", "Cédula", "Cat. Cliente", "Vendedor",
        "Subtotal", "Descuento", "IVA", "Total"
    ]
    _estilo_header(ws, 4, columnas)

    fmt_money = '#,##0.00'
    tot_sub = tot_desc = tot_iva = tot_total = 0

    for i, r in enumerate(datos, 5):
        bg = GRIS if i % 2 == 0 else BLANCO
        fill = PatternFill("solid", fgColor=bg)

        vals = [
            r.get("num_doc"), r.get("fecha"), r.get("tipo_documento"),
            r.get("tipo_venta"), r.get("cliente"), r.get("cedula"),
            r.get("cat_cliente"), r.get("vendedor"),
            r.get("subtotal", 0), r.get("descuento", 0),
            r.get("iva", 0), r.get("total", 0)
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = fill
            cell.border = _borde()
            cell.font   = Font(size=10)
            if col >= 9:
                cell.number_format = fmt_money
                cell.alignment = Alignment(horizontal="right")

        tot_sub   += r.get("subtotal", 0) or 0
        tot_desc  += r.get("descuento", 0) or 0
        tot_iva   += r.get("iva", 0) or 0
        tot_total += r.get("total", 0) or 0

    fila_tot = len(datos) + 5
    ws.merge_cells(f"A{fila_tot}:H{fila_tot}")
    c = ws.cell(row=fila_tot, column=1, value=f"TOTALES — {len(datos)} registros")
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=AZUL_CLAR)

    for col, val in enumerate([tot_sub, tot_desc, tot_iva, tot_total], 9):
        cell = ws.cell(row=fila_tot, column=col, value=val)
        cell.font          = Font(bold=True, size=10)
        cell.number_format = fmt_money
        cell.alignment     = Alignment(horizontal="right")
        cell.fill          = PatternFill("solid", fgColor=AZUL_CLAR)

    anchos = [10, 12, 12, 12, 30, 16, 14, 20, 14, 14, 14, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_pagos(datos: list, titulo: str = "Reporte de Pagos de Clientes") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"

    _titulo_reporte(ws, titulo)

    columnas = ["Nº Doc", "Fecha", "Cliente", "Cédula",
                "Tipo Pago", "Monto", "Saldo Ant.", "Saldo Act.", "Monto Aplicado"]
    _estilo_header(ws, 4, columnas)

    fmt_money = '#,##0.00'
    tot_monto = tot_aplicado = 0

    for i, r in enumerate(datos, 5):
        bg   = GRIS if i % 2 == 0 else BLANCO
        fill = PatternFill("solid", fgColor=bg)
        vals = [
            r.get("num_doc"), r.get("fecha"), r.get("cliente"),
            r.get("cedula"), r.get("tipo_pago"),
            r.get("monto", 0), r.get("saldo_anterior", 0),
            r.get("saldo_actual", 0), r.get("monto_aplicado", 0)
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = fill
            cell.border = _borde()
            cell.font   = Font(size=10)
            if col >= 6:
                cell.number_format = fmt_money
                cell.alignment = Alignment(horizontal="right")

        tot_monto    += r.get("monto", 0) or 0
        tot_aplicado += r.get("monto_aplicado", 0) or 0

    fila_tot = len(datos) + 5
    ws.merge_cells(f"A{fila_tot}:E{fila_tot}")
    c = ws.cell(row=fila_tot, column=1, value=f"TOTALES — {len(datos)} registros")
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=AZUL_CLAR)

    for col, val in enumerate([tot_monto, 0, 0, tot_aplicado], 6):
        cell = ws.cell(row=fila_tot, column=col, value=val)
        cell.font          = Font(bold=True, size=10)
        cell.number_format = fmt_money
        cell.alignment     = Alignment(horizontal="right")
        cell.fill          = PatternFill("solid", fgColor=AZUL_CLAR)

    anchos = [10, 12, 30, 16, 16, 14, 14, 14, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_productos(datos: list, titulo: str = "Productos Vendidos") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    _titulo_reporte(ws, titulo)

    columnas = ["Código", "Descripción", "Unidad",
                "Cantidad", "Subtotal", "Descuento", "IVA", "Total"]
    _estilo_header(ws, 4, columnas)

    fmt_money = '#,##0.00'
    fmt_cant  = '#,##0.00'
    tot_cant = tot_sub = tot_desc = tot_iva = tot_total = 0

    for i, r in enumerate(datos, 5):
        bg   = GRIS if i % 2 == 0 else BLANCO
        fill = PatternFill("solid", fgColor=bg)
        vals = [
            r.get("codigo"), r.get("descripcion"), r.get("unidad"),
            r.get("cantidad", 0), r.get("subtotal", 0),
            r.get("descuento", 0), r.get("iva", 0), r.get("total", 0),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = fill
            cell.border = _borde()
            cell.font   = Font(size=10)
            if col == 4:
                cell.number_format = fmt_cant
                cell.alignment = Alignment(horizontal="right")
            elif col >= 5:
                cell.number_format = fmt_money
                cell.alignment = Alignment(horizontal="right")

        tot_cant  += r.get("cantidad",  0) or 0
        tot_sub   += r.get("subtotal",  0) or 0
        tot_desc  += r.get("descuento", 0) or 0
        tot_iva   += r.get("iva",       0) or 0
        tot_total += r.get("total",     0) or 0

    fila_tot = len(datos) + 5
    ws.merge_cells(f"A{fila_tot}:C{fila_tot}")
    c = ws.cell(row=fila_tot, column=1, value=f"TOTALES — {len(datos)} productos")
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=AZUL_CLAR)

    for col, val in enumerate([tot_cant, tot_sub, tot_desc, tot_iva, tot_total], 4):
        cell = ws.cell(row=fila_tot, column=col, value=val)
        cell.font          = Font(bold=True, size=10)
        cell.number_format = fmt_money
        cell.alignment     = Alignment(horizontal="right")
        cell.fill          = PatternFill("solid", fgColor=AZUL_CLAR)

    anchos = [14, 40, 10, 12, 14, 14, 14, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_compras(datos: list, titulo: str = "Reporte de Compras") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    _titulo_reporte(ws, titulo)

    columnas = [
        "Nº Doc", "Fecha", "Proveedor", "Cédula",
        "Factura Prov.", "Subtotal", "Descuento", "IVA", "Total", "Estado"
    ]
    _estilo_header(ws, 4, columnas)

    fmt_money = '#,##0.00'
    tot_sub = tot_desc = tot_iva = tot_total = 0

    for i, r in enumerate(datos, 5):
        bg   = GRIS if i % 2 == 0 else BLANCO
        fill = PatternFill("solid", fgColor=bg)
        vals = [
            r.get("num_doc"), r.get("fecha"), r.get("proveedor"),
            r.get("cedula"), r.get("factura_proveedor"),
            r.get("subtotal", 0), r.get("descuento", 0),
            r.get("iva", 0), r.get("total", 0), r.get("estado")
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = fill
            cell.border = _borde()
            cell.font   = Font(size=10)
            if 6 <= col <= 9:
                cell.number_format = fmt_money
                cell.alignment = Alignment(horizontal="right")

        tot_sub   += r.get("subtotal", 0) or 0
        tot_desc  += r.get("descuento", 0) or 0
        tot_iva   += r.get("iva", 0) or 0
        tot_total += r.get("total", 0) or 0

    fila_tot = len(datos) + 5
    ws.merge_cells(f"A{fila_tot}:E{fila_tot}")
    c = ws.cell(row=fila_tot, column=1, value=f"TOTALES — {len(datos)} registros")
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=AZUL_CLAR)

    for col, val in enumerate([tot_sub, tot_desc, tot_iva, tot_total], 6):
        cell = ws.cell(row=fila_tot, column=col, value=val)
        cell.font          = Font(bold=True, size=10)
        cell.number_format = fmt_money
        cell.alignment     = Alignment(horizontal="right")
        cell.fill          = PatternFill("solid", fgColor=AZUL_CLAR)

    anchos = [10, 12, 30, 16, 16, 14, 14, 14, 14, 10]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_cxc(datos: list, titulo: str = "Cuentas por Cobrar") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CxC"

    _titulo_reporte(ws, titulo)

    columnas = ["Código", "Cliente", "Cédula", "Categoría", "Saldo"]
    _estilo_header(ws, 4, columnas)

    fmt_money = '#,##0.00'
    tot_saldo = 0

    for i, r in enumerate(datos, 5):
        bg   = GRIS if i % 2 == 0 else BLANCO
        fill = PatternFill("solid", fgColor=bg)
        vals = [
            r.get("codigo"),
            r.get("cliente"),
            r.get("cedula"),
            r.get("desc_categoria") or r.get("categoria"),
            r.get("saldo", 0),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = fill
            cell.border = _borde()
            cell.font   = Font(size=10)
            if col == 5:
                cell.number_format = fmt_money
                cell.alignment = Alignment(horizontal="right")

        tot_saldo += r.get("saldo", 0) or 0

    fila_tot = len(datos) + 5
    ws.merge_cells(f"A{fila_tot}:D{fila_tot}")
    c = ws.cell(row=fila_tot, column=1, value=f"TOTAL — {len(datos)} clientes")
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=AZUL_CLAR)

    cell = ws.cell(row=fila_tot, column=5, value=tot_saldo)
    cell.font          = Font(bold=True, size=10)
    cell.number_format = fmt_money
    cell.alignment     = Alignment(horizontal="right")
    cell.fill          = PatternFill("solid", fgColor=AZUL_CLAR)

    anchos = [10, 40, 18, 18, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
